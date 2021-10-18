# -*- coding: utf-8 -*-
import datetime
import os
import shutil
import winreg

from openpyxl import load_workbook, Workbook

class OutExcel():
    def __init__(self,outName):
        self.outName = outName

        self.copyFile()

    # 使用shutil完全复制文件，可以保持excel文件中表格的样式与模板的一致
    def copyFile(self):
        temp = shutil.copy('\\\\192.168.20.4/共享文件/资材部/原价程序文件/模板文件/模板文件.xlsx', self.outName)
        print(temp)
        if os.path.exists(self.outName):
            print("目标文件复制成功！")
        else:
            print("目标文件复制不成功！")

    # 输出Excel,建单界面使用
    def output_excel(self,fname,list1,list2,list3,list4):
        baozhuanbioa_db = list4 #  包装表
        jiagongbiao_db = list3  #  加工表
        cailiaobiao_db = list2  #  材料表
        zongbiao_db = list1 #  总表

        if not os.path.exists(fname):
            print(f"目标Excel文件不存在！")
            return

        wb = load_workbook(fname)
        ws = wb.active

        ws.title = str(zongbiao_db[1])  # 更改表名为 成品编码

        '''
        表头及表尾部分
        '''

        ws['B2'] = zongbiao_db[2]  # 客户名
        ws['B3'] = zongbiao_db[3]  # 品名
        ws['E2'] = zongbiao_db[14]  # 月产量
        ws['J1'] = zongbiao_db[23]  # 作成日期
        ws['J2'] = zongbiao_db[24]  # 作成
        ws['J3'] = zongbiao_db[25]  # 整理
        ws['L27'] = zongbiao_db[18]  # 材料费总额
        ws['L57'] = zongbiao_db[19]  # 加工费总额
        ws['L87'] = zongbiao_db[20]  # 包装费总额
        ws['G88'] = zongbiao_db[15]  # 固定费用
        ws['G89'] = zongbiao_db[16]  # 运输费
        ws['G90'] = zongbiao_db[17]  # 利润
        ws['C88'] = zongbiao_db[21]  # 总金额
        ws['D91'] = zongbiao_db[22]  # 生效日期
        ws['B93'] = zongbiao_db[26]  # 备注

        # 测试变量
        # print('包装表列表变量长度：', len(list4))
        # print('包装表列表变量值：', list4)
        # print('加工表列表变量长度：', len(list3))
        # print('加工表列表变量值：', list3)
        # print('材料表列表变量长度：', len(list2))
        # print('材料表列表变量值：', list2)

        # 定义变量
        bz_rows = len(list4)  # 获取包装数据的行数
        jg_rows = len(list3)  # 获取加工数据的行数
        cl_rows = len(list2)  # 获取材料数据的行数
        # print(type(bz_rows))
        # print('包装数据的行数为：',bz_rows)
        '''
        模板表格，默认值：
        材料费 20行  从 7  --- 26
        加工费 26行  从 31 --- 56
        包装费 26行  从 61 --- 86

        采用隐藏多余行的方法不显示空行

        '''

        # 包装费
        for x in range(bz_rows):
            ws[str('B' + str(61 + x))] = str(baozhuanbioa_db[x][4])  # 包装名称
            ws[str('D' + str(61 + x))] = str(baozhuanbioa_db[x][5])  # 单价(HKD)
            ws[str('F' + str(61 + x))] = str(baozhuanbioa_db[x][6])  # 用量
            ws[str('H' + str(61 + x))] = str(baozhuanbioa_db[x][7])  # 包装数
            ws[str('I' + str(61 + x))] = str(baozhuanbioa_db[x][8])  # 回收否
            ws[str('J' + str(61 + x))] = str(baozhuanbioa_db[x][9])  # 计算方式
            ws[str('L' + str(61 + x))] = str(baozhuanbioa_db[x][10])  # 金额
        # 隐藏多余的空行
        ws.row_dimensions.group(61 + bz_rows, 86, hidden=True)

        # 加工费
        for x in range(jg_rows):
            ws[str('B' + str(31 + x))] = str(jiagongbiao_db[x][3])  # 工序名称
            ws[str('D' + str(31 + x))] = str(jiagongbiao_db[x][6])  # 工数
            ws[str('F' + str(31 + x))] = str(jiagongbiao_db[x][5])  # 工程单价
            ws[str('H' + str(31 + x))] = str(jiagongbiao_db[x][7])  # 切割数
            ws[str('I' + str(31 + x))] = str(jiagongbiao_db[x][8])  # 用量
            ws[str('J' + str(31 + x))] = str(jiagongbiao_db[x][9])  # 计算方式
            ws[str('L' + str(31 + x))] = str(jiagongbiao_db[x][10])  # 金额

        ws.row_dimensions.group(31 + jg_rows, 56, hidden=True)

        # 材料费
        for x in range(cl_rows):
            ws[str('B' + str(7 + x))] = str(cailiaobiao_db[x][4])  # 名称
            ws[str('D' + str(7 + x))] = str(cailiaobiao_db[x][5])  # 供应商
            ws[str('F' + str(7 + x))] = str(cailiaobiao_db[x][6])  # 单价
            ws[str('G' + str(7 + x))] = str(cailiaobiao_db[x][7])  # 用量
            ws[str('H' + str(7 + x))] = str(cailiaobiao_db[x][8])  # 每模用量
            ws[str('I' + str(7 + x))] = str(cailiaobiao_db[x][9])  # 每模出数
            ws[str('J' + str(7 + x))] = str(cailiaobiao_db[x][10])  # 切割个数
            ws[str('L' + str(7 + x))] = str(cailiaobiao_db[x][11])  # 金额

        ws.row_dimensions.group(7 + cl_rows, 26, hidden=True)

        wb.save(fname)
        wb.close()
        os.startfile(fname)  # 调用系统默认程序打开，输出的Excel文件


#  把原价单列表导出到excel文件中
class OrderListExportToFile():
    def __init__(self):
        super(OrderListExportToFile, self).__init__()


    #  建立excel文件
    def createExcel(self):
        excel_name = f"{self.get_desktop()}\\{self.string_from_date()}.xlsx"
        print(f"文件名：{excel_name}")
        wb = Workbook(write_only=True)
        wb.save(excel_name)
        wb.close()
        return excel_name

    # 写excel文件
    def writeExcel(self,value_list):
        excel_name = self.createExcel()  # 创建excel文件,并获取文件名

        wb = load_workbook(excel_name)
        ws = wb.active

        for x in range(len(value_list)):
            ws.append(value_list[x])

        wb.save(excel_name)
        wb.close()

    # 获取当前电脑的桌面路径
    def get_desktop(self):
        # 利用系统的注册表
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                             r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
        # 返回的是Unicode类型数据
        return winreg.QueryValueEx(key, "Desktop")[0]

    # 按日期生成随机字符串
    def string_from_date(self):
        """按日期生成字符串用做成品编码(12位编码，类型为str，精确到分钟。例：202101220833)"""
        return '{0:%Y%m%d%H%M%S}'.format(datetime.datetime.now())